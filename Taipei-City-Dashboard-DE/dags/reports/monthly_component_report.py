"""
monthly_component_report
========================

每月 1 號自動產出「正在運作中的儀表板組件清單」Excel,
並 email 給承辦/相關人員。

資料來源:
    dashboardmanager DB
      - dashboard_groups (group_id 171=台北儀表板, 172=雙北儀表板)
      - dashboards (components 為 component_id 陣列)
      - components (id, index, name)
      - component_charts (依 index 取 chart 類型)
      - query_charts    (依 index 取 source / 更新頻率 / 加入日期)

產出:
    Sheet 1「運作中組件清單」明細
    Sheet 2「數量總覽」交叉表 + 計算口徑說明

需要事先在 Airflow 設定:
    Connection: dashboardmanager_pg  (PostgreSQL → dashboardmanager DB)
    Variable:   COMPONENT_REPORT_MAIL_LIST
                  例如 ["xxx@taipei.gov.tw", "yyy@example.com"]
                  也可在 default_args.email 直接列,或塞進 *MAIL_LIST 慣例
                  (CommonDag.fetch_email_list 風格,但本 DAG 沒走 CommonDag,
                   為避免散亂這支 DAG 直接讀單一 Variable)
"""
import io
import os
from collections import Counter
from datetime import datetime

from airflow import DAG
from airflow.models import Variable
from airflow.operators.python import PythonOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.utils.email import send_email


# ============================================================
#                    Configuration
# ============================================================
DAG_ID = "monthly_component_report"
PG_CONN_ID = "dashboardmanager_pg"
RECIPIENTS_VAR = "COMPONENT_REPORT_MAIL_LIST"
OUTPUT_DIR = "/tmp"

GROUP_TAIPEI = 171
GROUP_METROTAIPEI = 172

SQL = """
WITH active AS (
  SELECT DISTINCT
         unnest(d.components) AS component_id,
         g.group_id,
         d.id   AS dashboard_id,
         d.name AS dashboard_name
  FROM public.dashboard_groups g
  JOIN public.dashboards d ON d.id = g.dashboard_id
  WHERE g.group_id IN (%(g_tp)s, %(g_mt)s)
),
scope AS (
  SELECT component_id,
         string_agg(
           DISTINCT CASE group_id
                      WHEN %(g_tp)s THEN 'taipei'
                      WHEN %(g_mt)s THEN 'metrotaipei'
                    END,
           ',' ORDER BY CASE group_id
                          WHEN %(g_tp)s THEN 'taipei'
                          WHEN %(g_mt)s THEN 'metrotaipei'
                        END
         ) AS city_scope,
         string_agg(DISTINCT dashboard_name, ' | ' ORDER BY dashboard_name) AS dashboards
  FROM active
  GROUP BY component_id
),
qc_agg AS (
  SELECT index,
         string_agg(DISTINCT source, ' | ')          AS query_source,
         string_agg(DISTINCT city,   ',')            AS qc_cities,
         MAX(update_freq)                            AS update_freq,
         MAX(update_freq_unit)                       AS update_freq_unit,
         BOOL_OR(query_chart IS NOT NULL)            AS has_query_chart,
         MIN(created_at)                             AS first_created_at
  FROM public.query_charts
  GROUP BY index
)
SELECT c.id           AS component_id,
       c.index        AS component_index,
       c.name         AS component_name,
       cc.types       AS chart_type,
       qa.query_source,
       qa.update_freq,
       qa.update_freq_unit,
       qa.qc_cities   AS query_chart_city,
       qa.first_created_at,
       s.city_scope,
       s.dashboards,
       CASE WHEN qa.has_query_chart THEN 1 ELSE 0 END AS has_query_chart
FROM scope s
JOIN public.components c             ON c.id    = s.component_id
LEFT JOIN public.component_charts cc ON cc.index = c.index
LEFT JOIN qc_agg qa                  ON qa.index = c.index
ORDER BY s.city_scope, c.id;
"""


# ============================================================
#                    Helpers
# ============================================================
def _norm_scope(s: str) -> str:
    if not s:
        return s
    parts = set(s.split(","))
    has_tp = "taipei" in parts
    has_mt = "metrotaipei" in parts
    if has_tp and has_mt:
        return "雙北共用"
    if has_tp:
        return "台北"
    if has_mt:
        return "雙北"
    return s


def _chart_type_str(ct):
    if ct is None:
        return ""
    if isinstance(ct, list):
        return ",".join(ct)
    return str(ct)


def _fmt_dt(d):
    return d.strftime("%Y-%m-%d") if d else ""


# ============================================================
#                    Tasks
# ============================================================
def build_xlsx(**context):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hook = PostgresHook(postgres_conn_id=PG_CONN_ID)
    rows = hook.get_records(SQL, parameters={"g_tp": GROUP_TAIPEI, "g_mt": GROUP_METROTAIPEI})

    columns = [
        "component_id", "component_index", "component_name", "chart_type",
        "query_source", "update_freq", "update_freq_unit", "query_chart_city",
        "first_created_at", "city_scope", "dashboards", "has_query_chart",
    ]
    raw = [dict(zip(columns, r)) for r in rows]

    priority = {"雙北共用": 0, "台北": 1, "雙北": 2}
    data = []
    for r in raw:
        freq = ""
        if r.get("update_freq") is not None:
            freq = f"{r['update_freq']} {r.get('update_freq_unit') or ''}".strip()
        data.append({
            "component_id":     r["component_id"],
            "component_index":  r["component_index"],
            "component_name":   r["component_name"],
            "城市範圍":         _norm_scope(r["city_scope"]),
            "所屬儀表板":        r["dashboards"],
            "圖表類型":          _chart_type_str(r["chart_type"]),
            "資料來源":          r.get("query_source") or "",
            "更新頻率":          freq,
            "加入日期":          _fmt_dt(r.get("first_created_at")),
            "query_chart.city": r.get("query_chart_city") or "",
            "有 query_chart":    "是" if r["has_query_chart"] else "否",
        })
    data.sort(key=lambda x: (priority.get(x["城市範圍"], 99), x["component_id"]))

    wb = openpyxl.Workbook()
    headers = [
        "component_id", "component_index", "component_name",
        "城市範圍", "所屬儀表板", "圖表類型", "資料來源",
        "更新頻率", "加入日期", "query_chart.city", "有 query_chart",
    ]

    # Sheet 1
    ws = wb.active
    ws.title = "運作中組件清單"
    ws.append(headers)
    hdr_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = hdr_fill
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

    scope_color = {"雙北共用": "FFE699", "台北": "DDEBF7", "雙北": "E2EFDA"}
    for d in data:
        ws.append([d[h] for h in headers])
        r_idx = ws.max_row
        fill_color = scope_color.get(d["城市範圍"])
        if fill_color:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=col_idx).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
    ws.freeze_panes = "A2"
    widths = {
        "component_id": 12, "component_index": 28, "component_name": 24,
        "城市範圍": 12, "所屬儀表板": 50, "圖表類型": 16, "資料來源": 30,
        "更新頻率": 14, "加入日期": 14, "query_chart.city": 16, "有 query_chart": 14,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 18)
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2
    ws2 = wb.create_sheet("數量總覽")
    ws2["A1"] = f"資料抓取時間:{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws2["A1"].font = Font(italic=True, color="808080")
    ws2.cell(row=3, column=1, value="城市範圍").fill = hdr_fill
    ws2.cell(row=3, column=2, value="組件數").fill = hdr_fill
    ws2.cell(row=3, column=1).font = Font(bold=True, color="FFFFFF", size=12)
    ws2.cell(row=3, column=2).font = Font(bold=True, color="FFFFFF", size=12)

    cnt = Counter(d["城市範圍"] for d in data)
    display_order = ["台北", "雙北共用", "雙北"]
    total = sum(cnt.values())
    r = 4
    for label in display_order:
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=cnt.get(label, 0))
        fill_color = scope_color.get(label)
        if fill_color:
            for col in (1, 2):
                ws2.cell(row=r, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
        r += 1
    ws2.cell(row=r, column=1, value="總計 (去重)").font = Font(bold=True)
    ws2.cell(row=r, column=2, value=total).font = Font(bold=True)
    for col in (1, 2):
        ws2.cell(row=r, column=col).fill = PatternFill(
            start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
        )

    notes_r = r + 3
    ws2.cell(row=notes_r, column=1, value="計算口徑說明").font = Font(bold=True, size=12, color="305496")
    notes = [
        "・本清單以 dashboardmanager DB 即時資料為準(group_id 171=台北儀表板、172=雙北儀表板)。",
        "・每個 component 只列一次 — 同時掛在台北和雙北儀表板的會標為「雙北共用」。",
        "・「總計 (去重)」= 雙北儀表板實際運作中、互不重覆的組件數。",
        f"・若計算「台北版總組件數」= 台北 + 雙北共用 = {cnt.get('台北', 0) + cnt.get('雙北共用', 0)}",
        f"・若計算「雙北版總組件數」= 雙北 + 雙北共用 = {cnt.get('雙北', 0) + cnt.get('雙北共用', 0)}",
    ]
    for i, n in enumerate(notes):
        ws2.cell(row=notes_r + 1 + i, column=1, value=n)
        ws2.merge_cells(
            start_row=notes_r + 1 + i, start_column=1,
            end_row=notes_r + 1 + i, end_column=6,
        )
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    # Save
    ts = context["data_interval_end"].in_timezone("Asia/Taipei").strftime("%Y%m")
    out_path = os.path.join(OUTPUT_DIR, f"運作中組件清單_{ts}.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path} ({total} components)")

    summary = {
        "total":        total,
        "taipei":       cnt.get("台北", 0),
        "metrotaipei":  cnt.get("雙北", 0),
        "shared":       cnt.get("雙北共用", 0),
        "tp_full":      cnt.get("台北", 0) + cnt.get("雙北共用", 0),
        "mt_full":      cnt.get("雙北", 0) + cnt.get("雙北共用", 0),
        "out_path":     out_path,
        "ts":           ts,
    }
    return summary


def email_report(**context):
    summary = context["ti"].xcom_pull(task_ids="build_xlsx")
    out_path = summary["out_path"]
    ts = summary["ts"]

    recipients_raw = Variable.get(RECIPIENTS_VAR, default_var=None)
    if not recipients_raw:
        raise ValueError(
            f"Airflow Variable `{RECIPIENTS_VAR}` is not set. "
            f'Set it as a JSON list, e.g. ["a@x.com","b@x.com"].'
        )
    from ast import literal_eval
    recipients = literal_eval(recipients_raw)
    if isinstance(recipients, str):
        recipients = [recipients]

    subject = f"[城市儀表板] 運作中組件月報 {ts[:4]}-{ts[4:]}"
    html = f"""
    <p>各位先進好,</p>
    <p>附件為城市儀表板「運作中組件清單」<b>{ts[:4]}-{ts[4:]}</b> 月報。</p>
    <p>本月統計(資料抓取時間 {datetime.now().strftime('%Y-%m-%d %H:%M')}):</p>
    <ul>
      <li>全部組件 (去重):<b>{summary['total']}</b></li>
      <li>台北版總組件數 (含共用):<b>{summary['tp_full']}</b></li>
      <li>雙北版總組件數 (含共用):<b>{summary['mt_full']}</b></li>
      <li>純台北獨有:{summary['taipei']}</li>
      <li>純雙北獨有:{summary['metrotaipei']}</li>
      <li>雙北共用:{summary['shared']}</li>
    </ul>
    <p>明細請見附件 Sheet「運作中組件清單」,數量總覽請見 Sheet「數量總覽」。</p>
    <p style="color:#888;font-size:0.9em">
      (此信由 Airflow 自動產出 — DAG: {DAG_ID})
    </p>
    """
    send_email(
        to=recipients,
        subject=subject,
        html_content=html,
        files=[out_path],
    )
    print(f"Email sent to {recipients}")


# ============================================================
#                    DAG
# ============================================================
default_args = {
    "owner": "airflow",
    "email_on_retry": False,
    "email_on_failure": True,
    "retries": 1,
    "retry_delay": 600,
}

with DAG(
    dag_id=DAG_ID,
    description="每月匯出 dashboardmanager 上正在運作中的儀表板組件清單並寄信",
    start_date=datetime(2026, 4, 1),
    schedule="0 9 1 * *",     # 每月 1 號 09:00 (Airflow timezone, default UTC — 換算成台北時間自行調整)
    catchup=False,
    tags=["report", "city_dashboard", "monthly"],
    default_args=default_args,
) as dag:

    t_build = PythonOperator(
        task_id="build_xlsx",
        python_callable=build_xlsx,
    )

    t_email = PythonOperator(
        task_id="email_report",
        python_callable=email_report,
    )

    t_build >> t_email
